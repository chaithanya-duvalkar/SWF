# # src/main.py

# import json
# from parser import get_section_range
# from excel_writer import write_excel
# from validator import validate_range


# MAP_FILE = "input/app.map"
# OUTPUT_FILE = "output/memory_layout.xlsx"
# CONFIG_FILE = "config/sections.json"


# def main():

#     with open(CONFIG_FILE) as f:
#         config = json.load(f)

#     results = []

#     for section in config["sections"]:

#         ram_section = section["ram_section"]
#         ram_name = section["ram_name"]

#         start, end = get_section_range(MAP_FILE, ram_name)

#         status = validate_range(start, end)

#         print(f"{ram_name} → {start} - {end} ({status})")

#         results.append({
#             "RAM Section": ram_section,
#             "RAM Name": ram_name,
#             "Start Address": start,
#             "End Address": end,
#             "Status": status
#         })

#     write_excel(results, OUTPUT_FILE)

#     print("\nExcel generated successfully!")


# if __name__ == "__main__":
#     main()


import re
import sys
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =========================================================
# Detect MAP format
# =========================================================
def detect_format(map_file_path):
    with open(map_file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            if "memory region" in line.lower():
                return "hitech"
            if re.search(r'\|\s*0x[0-9A-Fa-f]+', line):
                return "ctc"
    return "unknown"

# =========================================================
# CTC Parser (Main + Nested)
# =========================================================
def parse_map_detailed_ctc(map_file_path):
    symbols = {}
    sizes = {}
    sections = []
    nested_sections = []

    with open(map_file_path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()

    # -------- Collect symbols --------
    for line in lines:
        match = re.search(r'\|\s*([A-Za-z0-9_]+)\s*\|\s*(0x[0-9A-Fa-f]+)', line)
        if match:
            symbols[match.group(1).upper()] = int(match.group(2), 16)

        size_match = re.search(r'([A-Za-z0-9_]+_SIZE)\s*\|\s*(0x[0-9A-Fa-f]+)', line)
        if size_match:
            sizes[size_match.group(1).upper()] = int(size_match.group(2), 16)

    # -------- Build sections --------
    for name, start_addr in symbols.items():
        if name.endswith("_START"):
            base = name.replace("_START", "")
            size = sizes.get(base + "_SIZE")
            end_addr = start_addr + size if size else None

            sections.append({
                "Section": base,
                "Start Address": hex(start_addr),
                "End Address": hex(end_addr) if end_addr else None,
                "Total Size (Hex)": hex(size) if size else None
            })

            # -------- Nested sections --------
            for sub_name, sub_start in symbols.items():
                if sub_name.startswith(base) and not sub_name.endswith("_START"):
                    sub_size = sizes.get(sub_name + "_SIZE")
                    sub_end = sub_start + sub_size if sub_size else None

                    status = "OK"
                    if end_addr and sub_end:
                        if sub_start < start_addr or sub_end > end_addr:
                            status = "OVERFLOW"

                    nested_sections.append({
                        "Parent Section": base,
                        "Sub-Region": sub_name,
                        "Start Address": hex(sub_start),
                        "End Address": hex(sub_end) if sub_end else None,
                        "Size (Hex)": hex(sub_size) if sub_size else None,
                        "Status": status
                    })

    return sections, nested_sections

# =========================================================
# Excel Formatting
# =========================================================
def format_excel(writer, sheet_name):
    ws = writer.book[sheet_name]

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Header formatting
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill

        ws.column_dimensions[get_column_letter(col)].width = 22

    # Status coloring
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=ws.max_column)
        if status_cell.value == "OK":
            status_cell.fill = ok_fill
        elif status_cell.value == "OVERFLOW":
            status_cell.fill = error_fill

# =========================================================
# Export Excel
# =========================================================
def export_to_excel(sections, nested_sections, output_file):
    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        pd.DataFrame(sections).to_excel(
            writer, sheet_name="Memory Sections", index=False
        )
        pd.DataFrame(nested_sections).to_excel(
            writer, sheet_name="Nested Sections", index=False
        )

        format_excel(writer, "Memory Sections")
        format_excel(writer, "Nested Sections")

    print(f"Excel created: {output_file}")
    print(f"Main sections  : {len(sections)}")
    print(f"Nested sections: {len(nested_sections)}")

# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python map_parser_ctc.py <map_file> [output.xlsx]")
        sys.exit(1)

    map_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "memory_sections.xlsx"

    fmt = detect_format(map_file)

    print("=" * 80)
    print("Enhanced MAP Parser with Nested + Reset-Safe Validation")
    print("=" * 80)

    if fmt == "ctc":
        sections, nested_sections = parse_map_detailed_ctc(map_file)
    else:
        print("Only CTC shown here (HiTech can be merged similarly)")
        sys.exit(1)

    export_to_excel(sections, nested_sections, output_file)
